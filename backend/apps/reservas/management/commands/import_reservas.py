import csv
import os
from django.core.management.base import BaseCommand
from apps.reservas.models import Reserva, Habitacion
from datetime import datetime
from typing import Dict, Any, List

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


def parse_float(value):
    # Elimina el signo de dólar y los puntos de los miles
    value = value.replace("$", "").replace(".", "")
    # Reemplaza la coma por el punto decimal
    value = value.replace(",", ".")
    return float(value)


class Command(BaseCommand):
    help = "Importar reservas desde archivo CSV o Excel (.xlsx)"

    def add_arguments(self, parser):
        parser.add_argument("path", type=str, help="Ruta a archivo .csv o .xlsx")

    def handle(self, *args, **options):
        path = options["path"]
        if not os.path.exists(path):
            self.stderr.write(self.style.ERROR(f"Archivo no encontrado: {path}"))
            return

        ext = os.path.splitext(path)[1].lower()

        try:
            if ext == ".csv":
                rows = self._read_csv(path)
            elif ext in (".xlsx", ".xlsm"):
                if load_workbook is None:
                    raise RuntimeError(
                        "openpyxl no instalado. Agrega 'openpyxl' a requirements.txt"
                    )
                rows = self._read_xlsx(path)
            else:
                self.stderr.write(
                    self.style.ERROR("Formato no soportado. Use .csv o .xlsx")
                )
                return
        except Exception as e:
            self.stderr.write(self.style.ERROR(f"Error leyendo archivo: {e}"))
            return

        resumen = {
            "procesadas": 0,
            "creadas": 0,
            "errores": 0,
            "habitaciones_creadas": 0,
            "detalles_error": [],
        }

        for row in rows:
            resumen["procesadas"] += 1
            try:
                data = self._normalize_row(row)

                habitacion, hab_created = Habitacion.objects.get_or_create(
                    numero=data["habitacion_numero"],
                    defaults={
                        "tipo": data.get("habitacion_tipo", "doble"),
                        "piso": data.get("habitacion_piso", "planta baja"),
                    },
                )
                if hab_created:
                    resumen["habitaciones_creadas"] += 1

                fecha_ingreso = self._parse_date(
                    data["check_in"]
                )  # dd/mm/YYYY o YYYY-MM-DD
                fecha_egreso = self._parse_date(data["check_out"])  # idem

                # Validar solapamiento de fechas en la misma habitación
                if Reserva.objects.filter(
                    nhabitacion=habitacion,
                    fecha_ingreso__lt=fecha_egreso,
                    fecha_egreso__gt=fecha_ingreso,
                ).exists():
                    raise ValueError(
                        f"Solapamiento: ya existe una reserva en {habitacion.numero} entre {fecha_ingreso} y {fecha_egreso}"
                    )

                # Evitar duplicado exacto (mismo huésped, mismas fechas y habitación)
                if Reserva.objects.filter(
                    nhabitacion=habitacion,
                    nombre=data["nombre"],
                    apellido=data["apellido"],
                    fecha_ingreso=fecha_ingreso,
                    fecha_egreso=fecha_egreso,
                ).exists():
                    raise ValueError(
                        f"Duplicada: {data['nombre']} {data['apellido']} en {habitacion.numero} con mismas fechas"
                    )

                Reserva.objects.create(
                    encargado=data.get("encargado", "Desconocido") or "Desconocido",
                    nhabitacion=habitacion,
                    nombre=data["nombre"],
                    apellido=data["apellido"],
                    personas=int(data.get("personas", 1) or 1),
                    fecha_ingreso=fecha_ingreso,
                    fecha_egreso=fecha_egreso,
                    noches=int(
                        data.get("noches")
                        or max((fecha_egreso - fecha_ingreso).days, 1)
                    ),
                    precio_por_noche=float(data.get("precio_por_noche") or 0.0),
                    monto_total=(
                        parse_float(str(data.get("monto_total", 0)))
                        if isinstance(data.get("monto_total"), str)
                        else float(data.get("monto_total", 0) or 0)
                    ),
                    senia=(
                        parse_float(str(data.get("senia", 0)))
                        if isinstance(data.get("senia"), str)
                        else float(data.get("senia", 0) or 0)
                    ),
                    resto=(
                        parse_float(str(data.get("resto", 0)))
                        if isinstance(data.get("resto"), str)
                        else float(data.get("resto", 0) or 0)
                    ),
                    cantidad_habitaciones=int(
                        data.get("cantidad_habitaciones", 1) or 1
                    ),
                    telefono=str(data.get("telefono", "")),
                    celiacos=self._parse_bool(data.get("celiacos")),
                    observaciones=str(data.get("observaciones", "")),
                    origen=str(data.get("origen", "")),
                )
                resumen["creadas"] += 1
            except Exception as e:
                resumen["errores"] += 1
                nombre = row.get("Nombre") or row.get("nombre") or "?"
                apellido = row.get("Apellido") or row.get("apellido") or "?"
                resumen["detalles_error"].append(f"{nombre} {apellido}: {e}")

        self.stdout.write(
            self.style.SUCCESS(
                f"Importación finalizada: procesadas={resumen['procesadas']}, creadas={resumen['creadas']}, errores={resumen['errores']}, habitaciones_creadas={resumen['habitaciones_creadas']}"
            )
        )
        if resumen["detalles_error"]:
            for d in resumen["detalles_error"]:
                self.stderr.write(self.style.WARNING(f"Error: {d}"))

    def _read_csv(self, path: str) -> List[Dict[str, Any]]:
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            return list(reader)

    def _read_xlsx(self, path: str) -> List[Dict[str, Any]]:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = {}
            for i, header in enumerate(headers):
                item[header] = row[i] if i < len(row) else None
            rows.append(item)
        return rows

    def _normalize_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        # Mapa de encabezados posibles -> clave estándar
        mapping = {
            "Habitación": "habitacion_numero",
            "Habitacion": "habitacion_numero",
            "habitación": "habitacion_numero",
            "Numero": "habitacion_numero",
            "Número": "habitacion_numero",
            "Tipo": "habitacion_tipo",
            "Piso": "habitacion_piso",
            "Check-In": "check_in",
            "Check In": "check_in",
            "Ingreso": "check_in",
            "Entrada": "check_in",
            "Check-Out": "check_out",
            "Check Out": "check_out",
            "Egreso": "check_out",
            "Salida": "check_out",
            "Nombre": "nombre",
            "Apellido": "apellido",
            "Personas": "personas",
            "Noches": "noches",
            "Precio por noche": "precio_por_noche",
            "Monto total": "monto_total",
            "Seña": "senia",
            "Senia": "senia",
            "Resto": "resto",
            "Cantidad\nde habitaciones": "cantidad_habitaciones",
            "Cantidad de habitaciones": "cantidad_habitaciones",
            "Telefono": "telefono",
            "Teléfono": "telefono",
            "Celiacos": "celiacos",
            "Observasiones": "observaciones",
            "Observaciones": "observaciones",
            "Origen": "origen",
            "Encargado": "encargado",
        }

        normalized: Dict[str, Any] = {}
        # Copiar tal cual por si ya vienen en formato estándar
        for k, v in row.items():
            key = mapping.get(str(k).strip(), str(k).strip())
            normalized[key] = v

        # Validaciones mínimas
        required = [
            "habitacion_numero",
            "nombre",
            "apellido",
            "check_in",
            "check_out",
            "monto_total",
            "senia",
            "origen",
        ]
        missing = [r for r in required if not normalized.get(r)]
        if missing:
            raise ValueError(
                f"Faltan columnas/valores requeridos: {', '.join(missing)}"
            )

        return normalized

    def _parse_date(self, value: Any) -> datetime.date:
        if value is None or value == "":
            raise ValueError("Fecha vacía")
        # Si viene como fecha de Excel (datetime/date), devolver date
        if hasattr(value, "date"):
            return value.date() if hasattr(value, "date") else value
        s = str(value).strip()
        # dd/mm/YYYY
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        raise ValueError(f"Formato de fecha inválido: {value}")

    def _parse_bool(self, value: Any) -> bool:
        if value is None:
            return False
        s = str(value).strip().lower()
        return s in ("si", "sí", "true", "1", "x", "s", "t")
