from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.authentication import SessionAuthentication
from django.contrib.auth import authenticate, login, logout
from django.db.models import Q, Sum, Count
from django.utils import timezone
from datetime import date, timedelta, datetime
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie
from django.middleware.csrf import get_token
from .models import Reserva, Habitacion, PerfilUsuario
from .serializers import (
    ReservaSerializer,
    HabitacionSerializer,
    UserSerializer,
    UserCreateUpdateSerializer,
    PerfilUsuarioSerializer,
    ReservaListSerializer,
    HabitacionListSerializer,
    EstadisticasSerializer,
    ReservaCreateSerializer,
    UsuarioListSerializer,
)
import csv
import os
import io
from typing import Any, Dict, List
from django.http import HttpResponse, Http404
from django.conf import settings
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.utils import ImageReader

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


class CsrfExemptSessionAuthentication(SessionAuthentication):
    """SessionAuthentication que no aplica verificación CSRF.

    Útil para endpoints específicos con multipart/form-data (uploads) en SPA.
    """

    def enforce_csrf(self, request):  # pragma: no cover
        return  # No-op: deshabilitar CSRF para esta autenticación


def is_supervisor(user):
    """Función helper para verificar si el usuario es supervisor"""
    try:
        # Verificar si el usuario es superusuario (tiene todos los permisos)
        if user.is_superuser:
            return True

        # Verificar si es staff (también tiene permisos de supervisor)
        if user.is_staff:
            return True

        # Verificar si tiene perfil y es supervisor
        if hasattr(user, "perfil") and user.perfil:
            return user.perfil.rol == "supervisor"

        # Si no tiene perfil, crear uno automáticamente
        try:
            perfil = PerfilUsuario.objects.get(usuario=user)
            return perfil.rol == "supervisor"
        except PerfilUsuario.DoesNotExist:
            # Crear perfil automáticamente
            perfil = PerfilUsuario.objects.create(
                usuario=user,
                rol=(
                    "supervisor" if (user.is_superuser or user.is_staff) else "conserge"
                ),
                turno="mañana",
                activo=True,
            )
            return perfil.rol == "supervisor"

        return False
    except Exception as e:
        print(f"Error verificando supervisor: {e}")
        return False


@method_decorator(ensure_csrf_cookie, name="get")
class AuthView(APIView):
    """Vista para autenticación"""

    authentication_classes = [SessionAuthentication]
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        """Login"""
        # Debug de CSRF/CORS: registrar el Origin y Host
        try:
            origin = request.META.get("HTTP_ORIGIN")
            host = request.get_host()
            print(f"[AuthView.post] Origin: {origin} | Host: {host}")
        except Exception:
            pass
        username = request.data.get("username")
        password = request.data.get("password")

        if not username or not password:
            return Response(
                {"error": "Usuario y contraseña son requeridos"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = authenticate(username=username, password=password)
        if user:
            login(request, user)
            serializer = UserSerializer(user)
            return Response({"user": serializer.data, "message": "Login exitoso"})
        else:
            return Response(
                {"error": "Credenciales inválidas"}, status=status.HTTP_401_UNAUTHORIZED
            )

    def delete(self, request):
        """Logout"""
        logout(request)
        return Response({"message": "Logout exitoso"})

    def get(self, request):
        """Obtener usuario actual"""
        if request.user.is_authenticated:
            serializer = UserSerializer(request.user)
            return Response(serializer.data)
        else:
            return Response(
                {"error": "No autenticado"}, status=status.HTTP_401_UNAUTHORIZED
            )


@method_decorator(ensure_csrf_cookie, name="dispatch")
class CsrfTokenView(APIView):
    """Devuelve el token CSRF para clientes SPA"""

    permission_classes = [permissions.AllowAny]

    def get(self, request):
        token = get_token(request)
        # Emitir clave estándar 'csrftoken' y setear cookie via ensure_csrf_cookie
        return Response({"csrftoken": token})


class FixAdminView(APIView):
    """Endpoint temporal para elevar permisos del admin sin shell"""

    permission_classes = [permissions.AllowAny]

    def post(self, request):
        # Solo funciona para usuario admin con password correcto
        username = request.data.get("username")
        password = request.data.get("password")

        if username != "admin":
            return Response({"error": "Solo para usuario admin"}, status=400)

        from django.contrib.auth import authenticate
        from django.contrib.auth.models import User

        user = authenticate(username=username, password=password)
        if not user:
            return Response({"error": "Credenciales incorrectas"}, status=401)

        # Elevar permisos
        user.is_staff = True
        user.is_superuser = True
        user.is_active = True
        user.save()

        # Asegurar perfil supervisor
        try:
            perfil, created = PerfilUsuario.objects.get_or_create(
                usuario=user,
                defaults={"rol": "supervisor", "turno": "mañana", "activo": True},
            )
            if not created and perfil.rol != "supervisor":
                perfil.rol = "supervisor"
                perfil.save()
        except Exception as e:
            pass

        return Response(
            {
                "message": "Admin elevado exitosamente",
                "is_staff": user.is_staff,
                "is_superuser": user.is_superuser,
                "rol": perfil.rol if "perfil" in locals() else "supervisor",
            }
        )


class HabitacionViewSet(viewsets.ModelViewSet):
    """ViewSet para el modelo Habitacion"""

    # Orden estable para evitar que elementos "desaparezcan" de la primera página al editar
    queryset = Habitacion.objects.all().order_by("id")
    serializer_class = HabitacionSerializer
    permission_classes = [permissions.IsAuthenticated]  # Requerir autenticación

    def get_serializer_class(self):
        if self.action == "list":
            return HabitacionListSerializer
        return HabitacionSerializer

    def create(self, request, *args, **kwargs):
        """Crear habitación - solo supervisores"""
        print(f"=== CREANDO HABITACIÓN ===")
        print(f"Usuario: {request.user.username}")
        print(f"Es superusuario: {request.user.is_superuser}")
        print(f"Es supervisor: {self._is_supervisor(request.user)}")

        if not self._is_supervisor(request.user):
            print(f"❌ Acceso denegado para {request.user.username}")
            return Response(
                {
                    "error": "Acceso denegado. Solo los supervisores pueden crear habitaciones."
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        print(f"✅ Acceso permitido para {request.user.username}")
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        """Actualizar habitación - solo supervisores"""
        if not self._is_supervisor(request.user):
            return Response(
                {
                    "error": "Acceso denegado. Solo los supervisores pueden editar habitaciones."
                },
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        """Eliminar habitación - solo supervisores"""
        if not self._is_supervisor(request.user):
            return Response(
                {
                    "error": "Acceso denegado. Solo los supervisores pueden eliminar habitaciones."
                },
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().destroy(request, *args, **kwargs)

    def _is_supervisor(self, user):
        """Verificar si el usuario es supervisor"""
        return is_supervisor(user)

    @action(detail=False, methods=["get"])
    def disponibles(self, request):
        """Obtener habitaciones disponibles"""
        # Por ahora retornamos todas las habitaciones ya que no hay campo estado
        habitaciones = Habitacion.objects.all()
        serializer = self.get_serializer(habitaciones, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def ocupadas(self, request):
        """Obtener habitaciones ocupadas"""
        # Por ahora retornamos todas las habitaciones ya que no hay campo estado
        habitaciones = Habitacion.objects.all()
        serializer = self.get_serializer(habitaciones, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def por_tipo(self, request):
        """Obtener habitaciones agrupadas por tipo"""
        tipo = request.query_params.get("tipo")
        if tipo:
            habitaciones = Habitacion.objects.filter(tipo=tipo)
        else:
            habitaciones = Habitacion.objects.all()

        serializer = self.get_serializer(habitaciones, many=True)
        return Response(serializer.data)


class ReservaViewSet(viewsets.ModelViewSet):
    """ViewSet para el modelo Reserva"""

    queryset = Reserva.objects.all().order_by("-id")
    serializer_class = ReservaSerializer
    permission_classes = [permissions.IsAuthenticated]  # Requerir autenticación

    def get_serializer_class(self):
        if self.action in [
            "list",
            "hoy",
            "checkins_hoy",
            "checkouts_hoy",
        ]:
            return ReservaListSerializer
        elif self.action == "create":
            return ReservaCreateSerializer
        return ReservaSerializer

    def update(self, request, *args, **kwargs):
        """Actualizar reserva - solo supervisores"""
        if not self._is_supervisor(request.user):
            return Response(
                {
                    "error": "Acceso denegado. Solo los supervisores pueden editar reservas."
                },
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        """Eliminar reserva - solo supervisores"""
        if not self._is_supervisor(request.user):
            return Response(
                {
                    "error": "Acceso denegado. Solo los supervisores pueden eliminar reservas."
                },
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().destroy(request, *args, **kwargs)

    def _is_supervisor(self, user):
        """Verificar si el usuario es supervisor"""
        return is_supervisor(user)

    def create(self, request, *args, **kwargs):
        """Crear reserva devolviendo el objeto completo con id"""
        create_serializer = ReservaCreateSerializer(data=request.data)
        create_serializer.is_valid(raise_exception=True)
        reserva = create_serializer.save()
        output_serializer = ReservaSerializer(reserva)
        headers = self.get_success_headers(output_serializer.data)
        return Response(
            output_serializer.data, status=status.HTTP_201_CREATED, headers=headers
        )

    @action(detail=True, methods=["get"], url_path="voucher")
    def voucher(self, request, pk=None):
        """Genera y devuelve un voucher PDF para la reserva"""
        try:
            reserva = Reserva.objects.select_related("nhabitacion").get(pk=pk)
        except Reserva.DoesNotExist:
            raise Http404("Reserva no encontrada")

        # Preparar respuesta PDF
        response = HttpResponse(content_type="application/pdf")
        filename = f"voucher_reserva_{reserva.id}.pdf"
        response["Content-Disposition"] = f'inline; filename="{filename}"'

        # Construcción del PDF con ReportLab (Platypus)
        doc = SimpleDocTemplate(
            response,
            pagesize=A4,
            leftMargin=18 * mm,
            rightMargin=18 * mm,
            topMargin=18 * mm,
            bottomMargin=18 * mm,
        )

        styles = getSampleStyleSheet()
        elements = []

        # Encabezado estilizado
        titulo = Paragraph(
            "<para align='center'><b>Voucher de Reserva</b></para>", styles["Title"]
        )
        subtitulo = Paragraph(
            f"<para align='center'>Reserva #{reserva.id} · Habitación {reserva.nhabitacion.numero} ({reserva.nhabitacion.tipo})</para>",
            styles["Normal"],
        )
        elements.extend([titulo, Spacer(1, 6), subtitulo, Spacer(1, 12)])

        # Callback para dibujar logo en esquina superior derecha
        def draw_header(canvas_obj, doc_obj):
            try:
                logo_path = os.path.join(
                    settings.BASE_DIR, "staticfiles", "images", "logo.png"
                )
                if os.path.exists(logo_path):
                    img = ImageReader(logo_path)
                    desired_width = 35 * mm
                    iw, ih = img.getSize()
                    scale = desired_width / float(iw)
                    desired_height = ih * scale
                    x = doc_obj.pagesize[0] - doc_obj.rightMargin - desired_width
                    y = doc_obj.pagesize[1] - doc_obj.topMargin - desired_height
                    canvas_obj.drawImage(
                        img,
                        x,
                        y,
                        width=desired_width,
                        height=desired_height,
                        preserveAspectRatio=True,
                        mask="auto",
                    )
            except Exception as e:
                # No interrumpir por errores de imagen
                pass

        # Datos del huésped
        huesped_data = [
            ["Nombre", f"{reserva.nombre} {reserva.apellido}"],
            ["Teléfono", reserva.telefono],
            ["Origen", reserva.origen],
            ["Encargado", reserva.encargado],
        ]

        fechas_data = [
            ["Fecha de ingreso", reserva.fecha_ingreso.strftime("%d/%m/%Y")],
            ["Fecha de salida", reserva.fecha_egreso.strftime("%d/%m/%Y")],
            ["Noches", str(reserva.noches)],
            ["Personas", str(reserva.personas)],
        ]

        montos_data = [
            ["Monto total", f"$ {reserva.monto_total:,.2f}".replace(",", ".")],
            ["Seña", f"$ {reserva.senia:,.2f}".replace(",", ".")],
            ["Resto", f"$ {reserva.resto:,.2f}".replace(",", ".")],
            [
                "Precio por noche",
                f"$ {reserva.precio_por_noche:,.2f}".replace(",", "."),
            ],
        ]

        def build_table(title: str, data: list) -> Table:
            header = Paragraph(f"<b>{title}</b>", styles["Heading3"])
            elements.extend([header, Spacer(1, 6)])
            table = Table(data, colWidths=[60 * mm, 100 * mm])
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 0), (-1, -1), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f9fafb")),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.HexColor("#ffffff"), colors.HexColor("#f3f4f6")],
                        ),
                        ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#e5e7eb")),
                        (
                            "INNERGRID",
                            (0, 0),
                            (-1, -1),
                            0.5,
                            colors.HexColor("#e5e7eb"),
                        ),
                    ]
                )
            )
            return table

        # Construir tablas
        elements.append(build_table("Datos del huésped", huesped_data))
        elements.append(Spacer(1, 10))
        elements.append(build_table("Estadía", fechas_data))
        elements.append(Spacer(1, 10))
        elements.append(build_table("Montos", montos_data))
        elements.append(Spacer(1, 14))

        # Notas importantes
        notas = [
            "Check-in a partir de las 12:00 del mediodía.",
            "Check-out a las 10:00 de la mañana.",
            "La totalidad de la reserva debe estar abonada para ingresar a la habitación.",
        ]
        for nota in notas:
            elements.append(
                Paragraph(f"<font color='#111827'>• {nota}</font>", styles["Normal"])
            )
            elements.append(Spacer(1, 4))

        # Observaciones
        if reserva.observaciones:
            elements.append(Spacer(1, 6))
            elements.append(Paragraph("<b>Observaciones</b>", styles["Heading3"]))
            elements.append(Spacer(1, 4))
            elements.append(Paragraph(reserva.observaciones, styles["Normal"]))

        # Renderizar
        doc.build(elements, onFirstPage=draw_header, onLaterPages=draw_header)

        return response

    @action(detail=False, methods=["get"])
    def hoy(self, request):
        """Obtener reservas de hoy"""
        hoy = date.today()
        reservas = Reserva.objects.filter(Q(fecha_ingreso=hoy) | Q(fecha_egreso=hoy))
        serializer = self.get_serializer(reservas, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def checkins_hoy(self, request):
        """Obtener check-ins de hoy"""
        hoy = date.today()
        reservas = Reserva.objects.filter(fecha_ingreso=hoy)
        serializer = self.get_serializer(reservas, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def checkouts_hoy(self, request):
        """Obtener check-outs de hoy"""
        hoy = date.today()
        reservas = Reserva.objects.filter(fecha_egreso=hoy)
        serializer = self.get_serializer(reservas, many=True)
        return Response(serializer.data)

    @action(
        detail=False,
        methods=["post"],
        url_path="importar",
        permission_classes=[permissions.IsAuthenticated],
        authentication_classes=[CsrfExemptSessionAuthentication],
    )
    def importar(self, request):
        """Importar reservas desde archivo CSV o Excel (.xlsx).

        Espera un campo 'file' en multipart/form-data.
        Crea habitaciones inexistentes automáticamente.
        Devuelve un resumen de la importación.
        """
        if not self._is_supervisor(request.user):
            return Response(
                {"error": "Acceso denegado. Solo supervisores pueden importar."},
                status=status.HTTP_403_FORBIDDEN,
            )

        uploaded = request.FILES.get("file")
        if not uploaded:
            return Response(
                {"error": "Archivo 'file' requerido"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        filename = getattr(uploaded, "name", "").lower()
        ext = os.path.splitext(filename)[1]

        try:
            if ext == ".csv" or uploaded.content_type in (
                "text/csv",
                "application/csv",
            ):
                data = uploaded.read().decode("utf-8", errors="ignore")
                rows = self._read_csv(io.StringIO(data))
            elif ext in (".xlsx", ".xlsm") or (
                uploaded.content_type
                in (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "application/vnd.ms-excel",
                )
            ):
                if load_workbook is None:
                    return Response(
                        {"error": "openpyxl no instalado en el servidor"},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    )
                rows = self._read_xlsx(uploaded)
            else:
                return Response(
                    {"error": "Formato no soportado. Use .csv o .xlsx"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        finally:
            try:
                uploaded.seek(0)
            except Exception:
                pass

        resumen = {
            "procesadas": 0,
            "creadas": 0,
            "errores": 0,
            "habitaciones_creadas": 0,
            "detalles_error": [],
        }

        for row in rows:
            resumen["procesadas"] += 1
            try:
                normalized = self._normalize_row(row)

                habitacion, hab_created = Habitacion.objects.get_or_create(
                    numero=normalized["habitacion_numero"],
                    defaults={
                        "tipo": normalized.get("habitacion_tipo", "doble"),
                        "piso": normalized.get("habitacion_piso", "planta baja"),
                    },
                )
                if hab_created:
                    resumen["habitaciones_creadas"] += 1

                fecha_ingreso = self._parse_date(
                    normalized["check_in"]
                )  # dd/mm/YYYY o YYYY-MM-DD
                fecha_egreso = self._parse_date(normalized["check_out"])  # idem

                # Validar solapamiento
                if Reserva.objects.filter(
                    nhabitacion=habitacion,
                    fecha_ingreso__lt=fecha_egreso,
                    fecha_egreso__gt=fecha_ingreso,
                ).exists():
                    raise ValueError(
                        f"Solapamiento: ya existe una reserva en {habitacion.numero} entre {fecha_ingreso} y {fecha_egreso}"
                    )

                # Duplicado exacto
                if Reserva.objects.filter(
                    nhabitacion=habitacion,
                    nombre=normalized["nombre"],
                    apellido=normalized["apellido"],
                    fecha_ingreso=fecha_ingreso,
                    fecha_egreso=fecha_egreso,
                ).exists():
                    raise ValueError(
                        f"Duplicada: {normalized['nombre']} {normalized['apellido']} en {habitacion.numero} con mismas fechas"
                    )

                # Crear reserva (cálculos automáticos en model.save)
                Reserva.objects.create(
                    encargado=normalized.get("encargado", "Desconocido")
                    or "Desconocido",
                    nhabitacion=habitacion,
                    nombre=normalized["nombre"],
                    apellido=normalized["apellido"],
                    personas=int(normalized.get("personas", 1) or 1),
                    fecha_ingreso=fecha_ingreso,
                    fecha_egreso=fecha_egreso,
                    monto_total=self._parse_number(normalized.get("monto_total", 0)),
                    senia=self._parse_number(normalized.get("senia", 0)),
                    resto=self._parse_number(normalized.get("resto", 0)),
                    cantidad_habitaciones=int(
                        normalized.get("cantidad_habitaciones", 1) or 1
                    ),
                    telefono=str(normalized.get("telefono", "")),
                    celiacos=self._parse_bool(normalized.get("celiacos")),
                    observaciones=str(normalized.get("observaciones", "")),
                    origen=str(normalized.get("origen", "")),
                )
                resumen["creadas"] += 1
            except Exception as e:
                resumen["errores"] += 1
                nombre = row.get("Nombre") or row.get("nombre") or "?"
                apellido = row.get("Apellido") or row.get("apellido") or "?"
                resumen["detalles_error"].append(f"{nombre} {apellido}: {e}")

        return Response(resumen)

    # ==== Helpers de importación ====
    def _read_csv(self, io_stream) -> List[Dict[str, Any]]:
        reader = csv.DictReader(io_stream)
        return list(reader)

    def _read_xlsx(self, file_obj) -> List[Dict[str, Any]]:
        wb = load_workbook(file_obj, data_only=True)
        ws = wb.active
        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        rows: List[Dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            item: Dict[str, Any] = {}
            for i, header in enumerate(headers):
                item[header] = row[i] if i < len(row) else None
            rows.append(item)
        return rows

    def _normalize_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        mapping = {
            "Habitación": "habitacion_numero",
            "Habitacion": "habitacion_numero",
            "habitación": "habitacion_numero",
            "Numero": "habitacion_numero",
            "Número": "habitacion_numero",
            "Tipo": "habitacion_tipo",
            "Piso": "habitacion_piso",
            "Check-In": "check_in",
            "Check In": "check_in",
            "Ingreso": "check_in",
            "Entrada": "check_in",
            "Check-Out": "check_out",
            "Check Out": "check_out",
            "Egreso": "check_out",
            "Salida": "check_out",
            "Nombre": "nombre",
            "Apellido": "apellido",
            "Personas": "personas",
            "Noches": "noches",
            "Precio por noche": "precio_por_noche",
            "Monto total": "monto_total",
            "Seña": "senia",
            "Senia": "senia",
            "Resto": "resto",
            "Cantidad\nde habitaciones": "cantidad_habitaciones",
            "Cantidad de habitaciones": "cantidad_habitaciones",
            "Telefono": "telefono",
            "Teléfono": "telefono",
            "Celiacos": "celiacos",
            "Observasiones": "observaciones",
            "Observaciones": "observaciones",
            "Origen": "origen",
            "Encargado": "encargado",
        }

        normalized: Dict[str, Any] = {}
        for k, v in row.items():
            key = mapping.get(str(k).strip(), str(k).strip())
            normalized[key] = v

        required = [
            "habitacion_numero",
            "nombre",
            "apellido",
            "check_in",
            "check_out",
            "monto_total",
            "senia",
            "origen",
        ]
        missing = [r for r in required if not normalized.get(r)]
        if missing:
            raise ValueError(
                f"Faltan columnas/valores requeridos: {', '.join(missing)}"
            )
        return normalized

    def _parse_date(self, value: Any) -> date:
        if value is None or value == "":
            raise ValueError("Fecha vacía")
        if hasattr(value, "date"):
            try:
                return value.date()
            except Exception:
                pass
        s = str(value).strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        raise ValueError(f"Formato de fecha inválido: {value}")

    def _parse_bool(self, value: Any) -> bool:
        if value is None:
            return False
        s = str(value).strip().lower()
        return s in ("si", "sí", "true", "1", "x", "s", "t")

    def _parse_number(self, value: Any) -> float:
        if value is None or value == "":
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value)
        s = s.replace("$", "").replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return 0.0

    @action(detail=False, methods=["get"])
    def por_fecha(self, request):
        """Obtener reservas por fecha específica"""
        # Debug: verificar autenticación
        print(f"=== DEBUG POR_FECHA ===")
        print(f"Usuario autenticado: {request.user.is_authenticated}")
        print(f"Usuario: {request.user}")
        print(f"Headers: {dict(request.headers)}")

        # Temporalmente permitir acceso sin autenticación para pruebas
        if not request.user.is_authenticated:
            print("Usuario no autenticado, pero continuando para pruebas...")

        fecha_str = request.query_params.get("fecha")
        if not fecha_str:
            return Response(
                {"error": "Parámetro fecha requerido"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            fecha = date.fromisoformat(fecha_str)
            reservas = Reserva.objects.filter(
                Q(fecha_ingreso__lte=fecha) & Q(fecha_egreso__gt=fecha)
            ).select_related(
                "nhabitacion"
            )  # Optimizar consulta

            # Calcular total de personas para el día actual
            total_personas_actual = (
                reservas.aggregate(total=Sum("personas"))["total"] or 0
            )

            # Cálculo de medialunas para el día siguiente basado en las personas del día actual
            medialunas_necesarias = (total_personas_actual * 2.5) / 12
            docenas_medialunas = round(medialunas_necesarias, 1)
            fecha_siguiente = fecha + timedelta(days=1)

            serializer = ReservaListSerializer(reservas, many=True)
            return Response(
                {
                    "reservas": serializer.data,
                    "medialunas": {
                        "fecha_siguiente": fecha_siguiente.isoformat(),
                        "total_personas": total_personas_actual,
                        "docenas_necesarias": docenas_medialunas,
                        "medialunas_totales": round(total_personas_actual * 2.5, 0),
                    },
                    "total_personas_actual": total_personas_actual,
                }
            )
        except ValueError:
            return Response(
                {"error": "Formato de fecha inválido (YYYY-MM-DD)"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=False, methods=["get"])
    def por_habitacion(self, request):
        """Obtener reservas por habitación"""
        habitacion_id = request.query_params.get("habitacion_id")
        if not habitacion_id:
            return Response(
                {"error": "Parámetro habitacion_id requerido"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            reservas = Reserva.objects.filter(nhabitacion_id=habitacion_id)
            serializer = self.get_serializer(reservas, many=True)
            return Response(serializer.data)
        except ValueError:
            return Response(
                {"error": "ID de habitación inválido"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=False, methods=["get"])
    def limpieza(self, request):
        """Obtener datos de limpieza para una fecha específica"""
        # Debug: verificar autenticación
        print(f"=== DEBUG LIMPIEZA ===")
        print(f"Usuario autenticado: {request.user.is_authenticated}")
        print(f"Usuario: {request.user}")
        print(f"Headers: {dict(request.headers)}")

        # Temporalmente permitir acceso sin autenticación para pruebas
        if not request.user.is_authenticated:
            print("Usuario no autenticado, pero continuando para pruebas...")

        fecha_str = request.GET.get("fecha")

        if not fecha_str:
            fecha_str = date.today().isoformat()

        try:
            fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        except ValueError:
            fecha = date.today()

        # Obtener todas las habitaciones
        habitaciones = Habitacion.objects.all()

        # Obtener reservas activas para la fecha
        reservas_activas = Reserva.objects.filter(
            fecha_ingreso__lte=fecha, fecha_egreso__gt=fecha
        ).select_related("nhabitacion")

        # Habitaciones ocupadas
        habitaciones_ocupadas = set(
            reservas_activas.values_list("nhabitacion_id", flat=True)
        )

        # Habitaciones a limpiar (se fueron los pasajeros)
        a_limpiar = []
        # Habitaciones a pasajero (está el pasajero pero hay que repasar)
        a_pasajero = []
        # Habitaciones a limpiar + pasajero (4ª noche, tiene al menos 1 noche más)
        a_limpiar_pasajero = []

        for habitacion in habitaciones:
            if habitacion.id in habitaciones_ocupadas:
                # Buscar la reserva activa para esta habitación
                reserva = reservas_activas.filter(nhabitacion=habitacion).first()

                if reserva:
                    # Calcular noches de estadía hasta la fecha
                    noches_estadia = (fecha - reserva.fecha_ingreso).days

                    # Si es la 4ª noche o más y tiene al menos 1 noche más
                    if (
                        noches_estadia >= 3
                        and reserva.fecha_egreso > fecha + timedelta(days=1)
                    ):
                        a_limpiar_pasajero.append(
                            {
                                "id": habitacion.id,
                                "numero": habitacion.numero,
                                "tipo": habitacion.tipo,
                                "noches_estadia": noches_estadia + 1,
                            }
                        )
                    else:
                        # Repaso normal
                        a_pasajero.append(
                            {
                                "id": habitacion.id,
                                "numero": habitacion.numero,
                                "tipo": habitacion.tipo,
                                "noches_estadia": noches_estadia + 1,
                            }
                        )
            else:
                # Habitación disponible - verificar si se fue alguien ayer
                reserva_ayer = Reserva.objects.filter(
                    nhabitacion=habitacion, fecha_egreso=fecha
                ).first()

                if reserva_ayer:
                    a_limpiar.append(
                        {
                            "id": habitacion.id,
                            "numero": habitacion.numero,
                            "tipo": habitacion.tipo,
                        }
                    )

        return Response(
            {
                "a_limpiar": a_limpiar,
                "a_pasajero": a_pasajero,
                "a_limpiar_pasajero": a_limpiar_pasajero,
                "fecha": fecha.isoformat(),
            }
        )


class EstadisticasView(APIView):
    """Vista para obtener estadísticas generales"""

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        hoy = date.today()

        # Estadísticas básicas
        total_reservas = Reserva.objects.count()
        total_habitaciones = Habitacion.objects.count()
        # Por ahora no hay campo estado, así que todas están disponibles
        habitaciones_ocupadas = 0
        habitaciones_disponibles = total_habitaciones

        # Ingresos totales
        ingresos_totales = (
            Reserva.objects.aggregate(total=Sum("monto_total"))["total"] or 0
        )

        # Reservas de hoy
        reservas_hoy = Reserva.objects.filter(fecha_ingreso=hoy).count()
        checkouts_hoy = Reserva.objects.filter(fecha_egreso=hoy).count()

        data = {
            "total_reservas": total_reservas,
            "total_habitaciones": total_habitaciones,
            "habitaciones_ocupadas": habitaciones_ocupadas,
            "habitaciones_disponibles": habitaciones_disponibles,
            "ingresos_totales": ingresos_totales,
            "reservas_hoy": reservas_hoy,
            "checkouts_hoy": checkouts_hoy,
        }

        serializer = EstadisticasSerializer(data)
        return Response(serializer.data)


class DashboardView(APIView):
    """Vista para datos del dashboard"""

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        hoy = date.today()

        # Últimas reservas
        ultimas_reservas = Reserva.objects.select_related("nhabitacion").order_by(
            "-id"
        )[:5]
        reservas_serializer = ReservaListSerializer(ultimas_reservas, many=True)

        # Check-ins de hoy
        checkins_hoy = Reserva.objects.filter(fecha_ingreso=hoy)
        checkins_serializer = ReservaListSerializer(checkins_hoy, many=True)

        # Check-outs de hoy
        checkouts_hoy = Reserva.objects.filter(fecha_egreso=hoy)
        checkouts_serializer = ReservaListSerializer(checkouts_hoy, many=True)

        # Habitaciones disponibles (todas por ahora)
        habitaciones_disponibles = Habitacion.objects.all()
        habitaciones_serializer = HabitacionListSerializer(
            habitaciones_disponibles, many=True
        )

        data = {
            "ultimas_reservas": reservas_serializer.data,
            "checkins_hoy": checkins_serializer.data,
            "checkouts_hoy": checkouts_serializer.data,
            "habitaciones_disponibles": habitaciones_serializer.data,
        }

        return Response(data)


class PerfilUsuarioViewSet(viewsets.ModelViewSet):
    """ViewSet para el modelo PerfilUsuario"""

    queryset = PerfilUsuario.objects.all()
    serializer_class = PerfilUsuarioSerializer
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=["get"], url_path="mi-perfil")
    def mi_perfil(self, request):
        """Obtener perfil del usuario actual"""
        try:
            perfil = PerfilUsuario.objects.get(usuario=request.user)
            serializer = self.get_serializer(perfil)
            return Response(serializer.data)
        except PerfilUsuario.DoesNotExist:
            # Crear perfil automáticamente si no existe
            perfil = PerfilUsuario.objects.create(
                usuario=request.user,
                rol=(
                    "supervisor"
                    if (request.user.is_superuser or request.user.is_staff)
                    else "conserge"
                ),
                turno="mañana",
                activo=True,
            )
            serializer = self.get_serializer(perfil)
            return Response(serializer.data)


class UsuarioViewSet(viewsets.ModelViewSet):
    """ViewSet para el modelo User"""

    from django.contrib.auth.models import User

    queryset = User.objects.all()
    permission_classes = [permissions.IsAuthenticated]  # Requerir autenticación

    def get_serializer_class(self):
        if self.action in ["create", "update", "partial_update"]:
            return UserCreateUpdateSerializer
        elif self.action == "list":
            return UsuarioListSerializer
        return UserSerializer

    def list(self, request, *args, **kwargs):
        """Listar usuarios con logs de depuración"""
        # Verificar si el usuario es supervisor
        if not self._is_supervisor(request.user):
            return Response(
                {
                    "error": "Acceso denegado. Solo los supervisores pueden ver usuarios."
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        print("=== LISTANDO USUARIOS ===")
        queryset = self.get_queryset()
        print(f"Queryset: {queryset}")
        print(f"Cantidad de usuarios: {queryset.count()}")

        serializer = self.get_serializer(queryset, many=True)
        print(f"Datos serializados: {serializer.data}")

        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        """Crear usuario - solo supervisores"""
        if not self._is_supervisor(request.user):
            return Response(
                {
                    "error": "Acceso denegado. Solo los supervisores pueden crear usuarios."
                },
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        """Actualizar usuario - solo supervisores"""
        if not self._is_supervisor(request.user):
            return Response(
                {
                    "error": "Acceso denegado. Solo los supervisores pueden editar usuarios."
                },
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        """Eliminar usuario - solo supervisores"""
        if not self._is_supervisor(request.user):
            return Response(
                {
                    "error": "Acceso denegado. Solo los supervisores pueden eliminar usuarios."
                },
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().destroy(request, *args, **kwargs)

    def _is_supervisor(self, user):
        """Verificar si el usuario es supervisor"""
        return is_supervisor(user)

    def perform_create(self, serializer):
        """Crear usuario con contraseña encriptada"""
        print("=== CREANDO USUARIO ===")
        user = serializer.save()
        print(f"Usuario creado: {user.username}")
        return user

    def perform_update(self, serializer):
        """Actualizar usuario"""
        user = serializer.save()
        return user


@method_decorator(csrf_exempt, name="dispatch")
class PlanningViewSet(viewsets.ViewSet):
    """ViewSet para el planning de reservas"""

    permission_classes = [permissions.AllowAny]  # Permitir acceso sin autenticación

    @action(detail=False, methods=["get"])
    def planning(self, request):
        """Obtener datos del planning de reservas"""
        start_date_str = request.GET.get("start_date")

        if start_date_str:
            try:
                first_day = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            except ValueError:
                first_day = date.today().replace(day=1)
        else:
            first_day = date.today().replace(day=1)

        # Generar 60 días desde la fecha de inicio
        days = [first_day + timedelta(days=i) for i in range(60)]

        # Obtener todas las habitaciones ordenadas por tipo
        habitaciones = list(Habitacion.objects.all())
        tipo_orden = {"doble": 1, "triple": 2, "cuadruple": 3, "quintuple": 4}
        habitaciones.sort(key=lambda x: tipo_orden.get(x.tipo, 5))

        # Obtener reservas que se superponen con el período
        reservas = Reserva.objects.filter(
            fecha_ingreso__lte=days[-1], fecha_egreso__gte=days[0]
        ).select_related("nhabitacion")

        planning_data = []
        for habitacion in habitaciones:
            ocupaciones = []
            nombre_mostrado = set()
            reservas_habitacion = reservas.filter(nhabitacion=habitacion)

            for day in days:
                ocupacion = None
                for reserva in reservas_habitacion:
                    if reserva.fecha_ingreso <= day < reserva.fecha_egreso:
                        if (
                            day == reserva.fecha_ingreso
                            and reserva.id not in nombre_mostrado
                        ):
                            nombre_mostrado.add(reserva.id)
                            ocupacion = {
                                "is_occupied": True,
                                "is_last_night": day
                                == reserva.fecha_egreso - timedelta(days=1),
                                "nombre": reserva.nombre,
                                "reserva_id": reserva.id,
                                "fecha_ingreso": reserva.fecha_ingreso.isoformat(),
                                "fecha_egreso": reserva.fecha_egreso.isoformat(),
                            }
                        else:
                            ocupacion = {
                                "is_occupied": True,
                                "is_last_night": day
                                == reserva.fecha_egreso - timedelta(days=1),
                                "nombre": None,
                                "reserva_id": reserva.id,
                                "fecha_ingreso": reserva.fecha_ingreso.isoformat(),
                                "fecha_egreso": reserva.fecha_egreso.isoformat(),
                            }
                        break

                if not ocupacion:
                    ocupacion = {
                        "is_occupied": False,
                        "is_last_night": False,
                        "nombre": None,
                        "reserva_id": None,
                        "fecha_ingreso": None,
                        "fecha_egreso": None,
                    }

                ocupaciones.append(ocupacion)

            planning_data.append(
                {
                    "habitacion": {
                        "id": habitacion.id,
                        "numero": habitacion.numero,
                        "tipo": habitacion.tipo,
                        "piso": habitacion.piso,
                    },
                    "ocupaciones": ocupaciones,
                }
            )

        return Response(
            {
                "planning": planning_data,
                "days": [day.isoformat() for day in days],
                "first_day": first_day.isoformat(),
            }
        )
